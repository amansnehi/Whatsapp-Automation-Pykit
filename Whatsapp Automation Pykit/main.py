import os
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import re


def send_message():

	contacts_name = list(map(str,input("Enter contacts each seperated by a \",\":  ").split(',')))
	text_message = input("Type Message: ")
	print(f"contacts list={contacts_name}")
	print(f"Your message = {text_message}")

	print("SIT BACK AND WATCH AUTOMATION DO IT WORK")

	#creating webdriver object
	driver = webdriver.Chrome("./chromedriver.exe", options=options)
	driver.maximize_window()
	driver.get("https://web.whatsapp.com")

	element = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.CLASS_NAME, "_3QfZd"))
	)  #wait until the content of the page load

	for contact in contacts_name:

		chat_button = driver.find_element_by_xpath("//span[@data-testid='chat'][@data-icon='chat']")
		chat_button.click()
		time.sleep(1)

		search_box = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='3']")
		search_box.click()
		time.sleep(2)

		search_box.send_keys(f"{contact}")
		time.sleep(1)

		try:

			selected_contact = driver.find_element_by_xpath("//span[@title='"+contact+"'][@class='_35k-1 _1adfa _3-8er']")
			selected_contact.click()

			time.sleep(2)
			type_message = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='6']")
			type_message.send_keys(f"{text_message}")
			time.sleep(2)
			driver.find_element_by_class_name("_1E0Oz").click()
			time.sleep(6)

		except:
			print(f"{contact} not found")
			back_button = driver.find_element_by_xpath("//span[@data-testid='back'][@data-icon='back']")
			back_button.click()
			time.sleep(1)


def bulk_message():
	#For accessing contacts name from Excel file
	wb_obj = openpyxl.load_workbook('./contacts.xlsx')
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	contacts_name = []

	for i in range(2,m_row+1):
		cell_obj = sheet_obj.cell(row=i,column=1)
		contacts_name.append(cell_obj.value)

	#to read a txt file (python provide function for creating, writing and reading txt files)
	file1 = open("text_message.txt","r+")
	text_message = file1.read().splitlines()

	print("SIT BACK AND WATCH AUTOMATION DO IT WORK")
	# print(f"contacts list={contacts_name}")
	# print(f"Your message = {text_message}")

	#creating webdriver object
	driver = webdriver.Chrome("./chromedriver.exe", options=options)
	driver.maximize_window()
	driver.get("https://web.whatsapp.com")

	element = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.CLASS_NAME, "_3QfZd"))
	)  #wait until the content of the page load

	for contact in contacts_name:

		chat_button = driver.find_element_by_xpath("//span[@data-testid='chat'][@data-icon='chat']")
		chat_button.click()
		time.sleep(1)

		search_box = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='3']")
		search_box.click()
		time.sleep(2)

		search_box.send_keys(f"{contact}")
		time.sleep(1)

		try:

			selected_contact = driver.find_element_by_xpath("//span[@title='"+contact+"'][@class='_35k-1 _1adfa _3-8er']")
			selected_contact.click()

			time.sleep(2)
			type_message = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='6']")
			for i in text_message:
				type_message.send_keys(i)
				ActionChains(driver).key_down(Keys.SHIFT).key_down(Keys.ENTER).key_up(Keys.SHIFT).key_up(Keys.ENTER).perform()
			time.sleep(2)
			driver.find_element_by_class_name("_1E0Oz").click()
			time.sleep(6)

		except:
			print(f"{contact} not found")
			back_button = driver.find_element_by_xpath("//span[@data-testid='back'][@data-icon='back']")
			back_button.click()
			time.sleep(1)


def forward_message(target_contact):

	# target_contact = input("Enter the contact name from where the messages is to be forwarded: ")
	print(f"target contact: {target_contact}")

	#For accessing contacts name from Excel file
	wb_obj = openpyxl.load_workbook('./groups.xlsx')
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	contacts_name = []

	for i in range(2,m_row+1):
		cell_obj = sheet_obj.cell(row=i,column=1)
		contacts_name.append(cell_obj.value)

	print("SIT BACK AND WATCH AUTOMATION DO IT WORK")
	
	#creating webdriver object
	driver = webdriver.Chrome("./chromedriver.exe", options=options)
	driver.maximize_window()
	driver.get("https://web.whatsapp.com")

	element = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.CLASS_NAME, "_3QfZd"))
	)  #wait until the content of the page loads

#repeat
	i = 0
	for _ in range(round(len(contacts_name)/5)):
		search_box = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='3']")
		search_box.click()
		time.sleep(2)

		search_box.send_keys(f"{target_contact}")
		time.sleep(1)

		selected_contact = driver.find_element_by_xpath("//span[@title='"+target_contact+"'][@class='_35k-1 _1adfa _3-8er']")
		selected_contact.click()
		time.sleep(2)

		forward_icon = driver.find_element_by_xpath("//span[@data-testid='forward-chat'][@data-icon='forward-chat']")
		forward_icon.click()
		time.sleep(2)

		search_box2 = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='3'][@dir='ltr']")
		# contacts = ['Work from Home #4','The Little Big Store','Daily only success trade','Easy Shopping','Motivational speaks','Work from Home #9']
		for contact in contacts_name[0+i:5+i]:
			search_box2.send_keys(f'{contact}')
			time.sleep(3)
			try:
				driver.find_element_by_class_name("-y4n1").click() #checkbox
				time.sleep(2)
				search_box2.clear()
			except:
				search_box2.clear()

		driver.find_element_by_class_name("SncVf._3doiV").click() #sendbutton
		i += 5
		time.sleep(3)


def links_share():

	#For accessing contacts name from Excel file
	wb_obj = openpyxl.load_workbook('./groups.xlsx')
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	contacts_name = []

	for i in range(2,m_row+1):
		cell_obj = sheet_obj.cell(row=i,column=1)
		contacts_name.append(cell_obj.value)

	#links-list
	wb_obj = openpyxl.load_workbook('./links.xlsx')
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	links_list = []

	for i in range(2,m_row+1):
		cell_obj = sheet_obj.cell(row=i,column=1)
		links_list.append(cell_obj.value)
	print("SIT BACK AND WATCH AUTOMATION DO IT WORK")
	# print(f"contacts list={contacts_name}")
	# print(f"Your message = {links_list}")
	
	#creating webdriver object
	driver = webdriver.Chrome("./chromedriver.exe", options=options)
	driver.maximize_window()
	driver.get("https://web.whatsapp.com")

	element = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.CLASS_NAME, "_3QfZd"))
	)  #wait until the content of the page load

	for contact in contacts_name:

		chat_button = driver.find_element_by_xpath("//span[@data-testid='chat'][@data-icon='chat']")
		chat_button.click()
		time.sleep(1)

		search_box = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='3']")
		search_box.click()
		time.sleep(2)

		search_box.send_keys(f"{contact}")
		time.sleep(1)

		try:
			selected_contact = driver.find_element_by_xpath("//span[@title='"+contact+"'][@class='_35k-1 _1adfa _3-8er']")
			selected_contact.click()

			time.sleep(2)
			type_message = driver.find_element_by_xpath("//div[@class='_2_1wd copyable-text selectable-text'][@contenteditable='true'][@data-tab='6']")
			for i in links_list:
				type_message.send_keys(i)
				time.sleep(4)
				driver.find_element_by_class_name("_1E0Oz").click()
			time.sleep(1)

		except:
			print(f"{contact} not found")
			back_button = driver.find_element_by_xpath("//span[@data-testid='back'][@data-icon='back']")
			back_button.click()
			time.sleep(1)


if __name__ == '__main__':

	dir_path = os.getcwd()        #to get the path of current working directory
	profile = os.path.join(dir_path, "profile", "wpp") #adding a new path profile/wpp folder in cwd 
	options = webdriver.ChromeOptions()  #to create an instance of chromeOptions
	options.add_argument(r"user-data-dir={}".format(profile)) #it will save the last session in the given path

	print("commands: send_message(), bulk_message(), forward_message(Enter the target contact), links_share()")

	# eval(input("Enter cmd: "))

	cmd_input = input("Enter cmd: ")

	lst = re.split('[(,)]',cmd_input)
	print(lst)
	cmd = lst[0]
	p1 = lst[1]
	p2 = lst[2]
	print(cmd,p1,p2)

	if p1 == '':
		eval(cmd + '()')
	elif p2 == '':
		eval(cmd + '(p1)')